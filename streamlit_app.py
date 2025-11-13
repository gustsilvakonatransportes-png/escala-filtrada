# streamlit_app.py
import streamlit as st
from io import BytesIO
from datetime import datetime
import pandas as pd
import re
from openpyxl import load_workbook

# Configuração da página
st.set_page_config(page_title="ESCALA FILTRADA", layout="wide")
st.title("🚛 Escala Filtrada — Extrator automático de blocos - José Cristiano")
st.markdown("Envie a planilha com os blocos; escolha o turno; baixe a planilha padronizada.")

# Escolha de turno
turno = st.selectbox("Escolha o turno:", ["Noturno", "Diurno"])

# Upload do arquivo Excel
uploaded_file = st.file_uploader("Enviar arquivo .xlsx (escala)", type=["xlsx"])

# 🔍 Padrões (regex)
FROTA_RE = re.compile(r"\b[TV]\d{2,4}\b", re.IGNORECASE)
PLACA_RE = re.compile(r"[A-Z0-9]{5,8}", re.IGNORECASE)
ROTA_RE = re.compile(r"\b\d{4,5}\b")
LARGADA_RE = re.compile(r"LARGAD|LARGA|LARGADA", re.IGNORECASE)
MOTORISTA_KEY = re.compile(r"motorista", re.IGNORECASE)
AJ1_KEY = re.compile(r"ajudante\s*1|aj1", re.IGNORECASE)
AJ2_KEY = re.compile(r"ajudante\s*2|aj2", re.IGNORECASE)

def cell_text(cell):
    return str(cell).strip() if cell else ""

def find_frota_lines(ws):
    positions = []
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = cell_text(ws.cell(row=r, column=c).value)
            if val and FROTA_RE.search(val):
                placa_candidate = ""
                for cc in range(c + 1, min(c + 4, max_col + 1)):
                    txt = cell_text(ws.cell(row=r, column=cc).value)
                    if txt and any(ch.isalpha() for ch in txt) and any(ch.isdigit() for ch in txt):
                        placa_candidate = txt
                        break
                positions.append((r, c, val, placa_candidate))
    return positions

def extract_blocks_by_frota(ws, frota_positions):
    blocks = []
    max_row = ws.max_row
    frota_positions = sorted(frota_positions, key=lambda x: (x[0], x[1]))
    if not frota_positions:
        blocks.append((1, max_row, 1, ws.max_column))
        return blocks
    for i, pos in enumerate(frota_positions):
        start_row = pos[0]
        end_row = frota_positions[i + 1][0] - 1 if i + 1 < len(frota_positions) else max_row
        blocks.append((start_row, end_row, 1, ws.max_column))
    return blocks

def extract_from_block(ws, start_row, end_row, start_col, end_col):
    frota = placa = rota = motorista = ajud1 = ajud2 = largada = ""
    for c in range(start_col, end_col + 1):
        v = cell_text(ws.cell(row=start_row, column=c).value)
        if FROTA_RE.search(v):
            frota = v
            for cc in range(c + 1, min(c + 4, end_col + 1)):
                txt = cell_text(ws.cell(row=start_row, column=cc).value)
                if txt and any(ch.isalpha() for ch in txt) and any(ch.isdigit() for ch in txt):
                    placa = txt
                    break
            break

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            v = cell_text(ws.cell(row=r, column=c).value)
            if v and ROTA_RE.search(v):
                rota = ROTA_RE.search(v).group(0).zfill(5)
                break
        if rota:
            break

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            v = cell_text(ws.cell(row=r, column=c).value)
            if not v:
                continue
            if MOTORISTA_KEY.search(v):
                motorista = cell_text(ws.cell(row=r + 1, column=c).value) or motorista
            if AJ1_KEY.search(v):
                ajud1 = cell_text(ws.cell(row=r + 1, column=c).value) or ajud1
            if AJ2_KEY.search(v):
                ajud2 = cell_text(ws.cell(row=r + 1, column=c).value) or ajud2
            if LARGADA_RE.search(v):
                mtime = re.search(r"(\d{1,2}[:h]\d{2})", v)
                if mtime:
                    largada = mtime.group(0).replace("h", ":")
                else:
                    largada = cell_text(ws.cell(row=r + 1, column=c).value) or largada

    return {
        "Frota": frota,
        "Placa": placa,
        "Rota": rota,
        "Motorista": motorista,
        "Ajudante 1": ajud1,
        "Ajudante 2": ajud2,
        "Largada": largada
    }

def parse_workbook_bytes(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    frota_positions = find_frota_lines(ws)
    blocks = extract_blocks_by_frota(ws, frota_positions)
    rows = []
    for b in blocks:
        data = extract_from_block(ws, *b)
        if any(data.values()):
            rows.append(data)
    df = pd.DataFrame(rows, columns=["Frota", "Placa", "Rota", "Motorista", "Ajudante 1", "Ajudante 2", "Largada"])
    return df

if uploaded_file:
    try:
        file_bytes = uploaded_file.read()
        df_out = parse_workbook_bytes(file_bytes)

        if df_out.empty:
            st.error("Não foi possível detectar blocos. Verifique o arquivo.")
        else:
            df_out["Turno"] = turno
            df_out = df_out[["Frota", "Placa", "Rota", "Motorista", "Ajudante 1", "Ajudante 2", "Turno", "Largada"]]

            st.success(f"✅ {len(df_out)} blocos encontrados!")

            # Cabeçalho em negrito com estilo garantido
            st.markdown("""
                <style>
                [data-testid="stTable"] thead tr th div p {
                    font-weight: 900 !important;
                    color: white !important;
                }
                </style>
            """, unsafe_allow_html=True)

            # Usar data_editor (modo leitura) — permite estilo
            st.data_editor(df_out, use_container_width=True, disabled=True)

            # Download do Excel
            buf = BytesIO()
            data_hoje = datetime.now().strftime("%d-%m-%Y")
            nome_arquivo = f"ESCALA_FILTRADA_{data_hoje}.xlsx"

            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_out.to_excel(writer, index=False, sheet_name="Escala Filtrada")

            st.download_button(
                "📥 Baixar ESCALA FILTRADA",
                buf.getvalue(),
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"❌ Erro ao processar a planilha: {e}")
